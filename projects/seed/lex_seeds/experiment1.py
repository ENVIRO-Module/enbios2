"""
first experiment
"""
from logging import getLogger
import pandas as pd
from collections import Counter
from enbios2.base.experiment import Experiment
import bw2data as bd
import bw2io
from projects.seed.Data.const import data_path
import openpyxl
import json
from bw2data.errors import UnknownObject
from bw2data.backends import Activity



#getLogger("peewee").setLevel("ERROR")

pass

processors_path=data_path / 'base_file_simplified.xlsx'
calliope=data_path / 'flow_out_sum_modified.csv'
dict_path=data_path /'dict.json'



data=openpyxl.load_workbook(processors_path)
processors=data.active
bd.projects.set_current('Hydrogen_SEEDS_3')
database=bd.Database('CUTOFF_391')

print(bd.projects.dir)
pass

processors=pd.read_excel(processors_path, sheet_name='BareProcessors simulation')

activities_cool={}
for index,row in processors.iterrows():
    code=str(row['BW_DB_FILENAME'])
    print(code)
    try:
        act=database.get(code)
    except UnknownObject:
        pass

    name=act['name']
    alias=row['Processor']+'_'+row['@SimulationCarrier']
    activities_cool[name]={
        'alias': alias,
        'code': code

    }

pass

"""

# collect processor, carrier and CODE
processors_data = []


for row in processors.iter_rows(min_row=2):
    if row[1].value:
        processors_data.append((row[1].value, row[2].value, row[20].value))
# JUST TAKE THE CODES
codes = [x[2] for x in processors_data]
# COPY, IN ORDER TO CHECK WHICH ARE MISSING
pass


pass
# read all values in column I

activities = []

for pd in processors_data:
    try:
        activity = database.get_node(pd[2])
        activities.append((activity, *pd))
    except UnknownObject:
        print(pd, "not found")
activities
pass


#redo the alias

potential_aliases = [act[1] for act in activities]

counts=Counter(potential_aliases)
print("All unique", len(potential_aliases) == len(set(potential_aliases)))

activities_ = []
for act in activities:
    full_alias = f"{act[1]}_{act[2]}"
    activities_.append((full_alias, *act[1:]))





activities_bw: dict[str, Activity] = {}
for activity in activities:
    #activities_bw[activity[0]] = database.get_node(activity[3])
    activities_bw[activity[1]] = database.get_node(activity[3])
    #alias=activities_bw[activity[1]]


activities_bw
pass
"""



SEEDS_DB_NAME = "seeds"
seeds_db = bd.Database(SEEDS_DB_NAME)

print(list(bd.databases))

SEEDS_DB_NAME = "seeds"

seeds_db = bd.Database(SEEDS_DB_NAME)
seeds_db.register()




from tqdm import tqdm

def full_duplicate(activity: Activity, code=None, **kwargs) -> Activity:
    """
    Make a copy of an activity with its upstream exchanges
    (Otherwise, you cannot calculate the lca of the copy)
    :param activity: the activity to copy
    :param code: code of the new activity
    :param kwargs: other data for the copy
    :return: new activity
    """
    activity_copy = activity.copy(code, **kwargs)
    for upstream in activity.upstream():
        upstream.output.new_exchange(input=activity_copy, type=upstream["type"], amount=upstream.amount).save()
    activity_copy.save()
    return activity_copy


for key,item in tqdm(activities_cool.items()):
    print(key)
    act=database.get(item['code'])
    alias=item['alias']
    #check if the activity is in the db:
    clone1 = full_duplicate(act)
    if clone1 in seeds_db:
        print('hey, I here',clone1['name'])


    else:
        clone1["name"] = clone1['name'] + "__PRT_1"
        clone1["database"] = SEEDS_DB_NAME
        clone1["alias"] = alias + "__PRT_1"
        clone1.save()


    clone2 = full_duplicate(act)
    if clone2 in seeds_db:
        pass
    else:
        clone2["name"] = clone2["name"] + "__PRT_2"
        clone2["database"] = SEEDS_DB_NAME
        clone2["alias"] = alias + "__PRT_2"
        clone2.save()
print(list(seeds_db))



pass



enbios2_activities = {
    activity["alias"]: {
        "id": {"code": activity["code"]}
    }
    for activity in list(seeds_db)
}
pass

from csv import DictReader
from enbios2.generic.files import ReadDataPath

enbios2scenarios = []
raw_scenarios = ReadDataPath(calliope) # .read_data()
raw_scenario_data = list(DictReader(raw_scenarios.open(encoding="utf-8"), delimiter=","))
pass
print(raw_scenario_data[:5])

# Scenarios_cool

def get_stuff(df):
    """

    :param df:
    :return:
    """
    stuff={}

    for index,row in df.iterrows():
        other_stuff=[]
        alias=row['aliases']
        flow_out_sum=(row['flow_out_sum'])
        unit=row['units']
        other_stuff.append(unit)
        other_stuff.append(flow_out_sum)

        stuff[alias]=other_stuff
    return stuff

def generate_scenarios(calliope_data,smaller_vers=False):


    cal_dat=pd.read_csv(calliope,delimiter=',')
    cal_dat['aliases']=cal_dat['techs'] + '_' + cal_dat['carriers'] + '__' + cal_dat['locs']
    scenarios=cal_dat['scenarios'].unique().tolist()
    if smaller_vers: # get a small version of the data
        scenarios=scenarios[:3]

    cooler_dooper={}

    for scenario in scenarios:
        df=cal_dat[cal_dat['scenarios']==scenario]
        stuff=get_stuff(df)
        cooler_dooper[scenario]={}
        cooler_dooper[scenario]['activities']=stuff


    return cooler_dooper



enbios2scenarios=generate_scenarios(calliope,smaller_vers=True)

""""

current_scenario_index = 0
current_scenario = {}
for row in raw_scenario_data:
    scen=str(row['scenarios']).split('_')[-1]
    scenario_index = int(scen)
    assert current_scenario_index <= scenario_index <= current_scenario_index + 1
    if scenario_index == current_scenario_index + 1:
        current_scenario_index = scenario_index
        enbios2scenarios.append({"activities": current_scenario.copy()})
        current_scenario = {}
    alias = f'{row["techs"]}_{row["carriers"]}__{row["locs"]}'
    current_scenario[alias] = [row["units"], row["flow_out_sum"]]

enbios2scenarios[0]
"""




from ast import literal_eval

#2=openpyxl.load_workbook(processors_path)

methods = data["ScalarIndicators"]

enbios2_methods = {

}

for row in methods.iter_rows(min_row=2):
    if row[5].value:
        method_tuple = literal_eval(row[5].value)
        enbios2_methods[method_tuple[2]] = {
            "id": method_tuple
        }

enbios2_methods={
'ozone depletion potential (ODPinfinite)': ('ReCiPe 2016 v1.03, midpoint (H)',
  'ozone depletion',
  'ozone depletion potential (ODPinfinite)'),
 'agricultural land occupation (LOP)': ('ReCiPe 2016 v1.03, midpoint (H)',
  'land use',
  'agricultural land occupation (LOP)'),
 'surplus ore potential (SOP)': ('ReCiPe 2016 v1.03, midpoint (H)',
  'material resources: metals/minerals',
  'surplus ore potential (SOP)'),
 'global warming potential (GWP1000)': ('ReCiPe 2016 v1.03, midpoint (H)',
  'climate change',
  'global warming potential (GWP1000)')
}
pass


enbios2_data = {
    "bw_project":'Hydrogen_SEEDS_3',
    "activities": enbios2_activities,
    "methods": enbios2_methods,
    "scenarios": enbios2scenarios
}

with open(dict_path, 'w') as gen_dict:
    json.dump(enbios2_data,gen_dict, indent=4)





from enbios2.base.experiment import Experiment
from enbios2.models.experiment_models import ExperimentData

exp_data = ExperimentData(**enbios2_data)
exp = Experiment(exp_data)

from pathlib import Path

Path("results").mkdir(exist_ok=True)

for scenario in exp.scenarios:
    print(scenario.alias)
    result_tree = scenario.run()
    scenario.results_to_csv(Path(f"results/{scenario.alias}.csv"), True)




pass


