"""
Alternative if no need to copy activities in a new DB
"""

import json
import random
from logging import getLogger

import bw2data
import bw2data as bd
import openpyxl
import pandas as pd
from bw2data.backends import Activity
from bw2data.errors import UnknownObject

from enbios2.generic.files import DataPath

getLogger("peewee").setLevel("ERROR")

data_path = DataPath("temp/seeds")
processors_path = data_path / 'base_file_simplified.xlsx'
calliope = data_path / 'flow_out_sum_modified.csv'
dict_path = data_path / 'enbios_input_2.json'

data = openpyxl.load_workbook(processors_path)
processors = data.active
bd.projects.set_current('ecoinvent')
database = bd.Database('cutoff_3.9.1_default')

print(bd.projects.dir)




# Scenarios_cool

def get_scenario(df):
    """
    Iters through 1 scenario of the data.csv (scenarios data), storing basic data in a dictionary
    Gets like
    activities : {
        alias : [
        unit,
        amount]
    }

    :param df:
    :return:
    """
    scenario = {}


    for index, row in df.iterrows():
        other_stuff = []
        alias = row['aliases']
        flow_out_sum = (row['flow_out_sum'])
        unit = row['units']
        other_stuff.append(unit)
        other_stuff.append(flow_out_sum)

        scenario[alias] = other_stuff
    return scenario




def generate_scenarios(calliope_data, smaller_vers=False):
    """
    #Iterate through all the data from calliope (data.csv, output results...)
        -Create new columns, such as alias
        -
    :param calliope_data:
    :param smaller_vers: BOOL, if true, a small version of the data for testing gets produced
    :return:
    """


    cal_dat = pd.read_csv(calliope_data, delimiter=',')
    cal_dat['aliases'] = cal_dat['techs'] + '_' + cal_dat['carriers'] + '___' + cal_dat['locs'] # Use ___ to split the loc for the recognision of the activities
    scenarios = cal_dat['scenarios'].unique().tolist()
    if smaller_vers:  # get a small version of the data
        scenarios = scenarios[:3]

    cooler_dooper = {}

    for scenario in scenarios:
        df = cal_dat[cal_dat['scenarios'] == scenario]
        stuff = get_scenario(df)
        cooler_dooper[scenario] = {}
        cooler_dooper[scenario]['activities'] = stuff

    # GENERATE KEYS FOR THE SCENARIOS

    scens = random.choice(list(cooler_dooper.keys()))  # select a random scenario from the list
    print(f'techs from scenario {scens} chosen')
    acts=list(cooler_dooper[scens]['activities'].keys())

    return cooler_dooper,acts


def generate_activities(*args):
    processors = pd.read_excel(processors_path, sheet_name='BareProcessors simulation')

    activities_cool = {}
    for index, row in processors.iterrows():
        code = str(row['BW_DB_FILENAME'])

        print('im parsing', str(row['Processor']), code)
        try:
            act = database.get_node(code)

        except UnknownObject:
            print(row['Processor'], 'has an unknown object')
            pass

        name = act['name']
        unit = act['unit']
        alias = str(row['Processor']) + '_' + str(row['@SimulationCarrier'])
        print(alias)

        activities_cool[alias] = {
            'name': name,
            'code': code,
        }

    # CALL THE FUNCTION HERE
    activities={}
    for element  in args:
        new_element=element.split('___')[0] #This should match the name
        for key in activities_cool.keys():
            if new_element==key:

                new_code=activities_cool[key]['code']
                activities[element]={
                    "id":{
                        'code':new_code
                    }

                }
    return activities



enbios2scenarios,activ_names = generate_scenarios(calliope, smaller_vers=True)

hope_final_acts=generate_activities(*activ_names)

from ast import literal_eval

# 2=openpyxl.load_workbook(processors_path)

methods = data["ScalarIndicators"]

enbios2_methods = {

}

for row in methods.iter_rows(min_row=2):
    if row[5].value:
        method_tuple = literal_eval(row[5].value)
        enbios2_methods[method_tuple[2]] = {
            "id": method_tuple
        }

enbios2_methods = {
    'ozone depletion potential (ODPinfinite)': ('ReCiPe 2016 v1.03, midpoint (H)',
                                                'ozone depletion',
                                                'ozone depletion potential (ODPinfinite)'),
    'agricultural land occupation (LOP)': ('ReCiPe 2016 v1.03, midpoint (H)',
                                           'land use',
                                           'agricultural land occupation (LOP)'),
    'surplus ore potential (SOP)': ('ReCiPe 2016 v1.03, midpoint (H)',
                                    'material resources: metals/minerals',
                                    'surplus ore potential (SOP)'),
    'global warming potential (GWP1000)': ('ReCiPe 2016 v1.03, midpoint (H)',
                                           'climate change',
                                           'global warming potential (GWP1000)')
}
pass

enbios2_data = {
    "bw_project": 'ecoinvent',
    "activities": hope_final_acts,
    "methods": enbios2_methods,
    "scenarios": enbios2scenarios
}

with open(dict_path, 'w') as gen_dict:
    json.dump(enbios2_data, gen_dict, indent=4)


# We're assuming that one scenario includes all the possible technologies


pass